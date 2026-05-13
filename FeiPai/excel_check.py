import openpyxl


class check_excel():

    def __init__(self):
        self.src_path = "src/game_excel/card_all.xlsx"
        self.wb = openpyxl.load_workbook(self.src_path, keep_vba=True)

    def get_feature_info(self):
        # wb = openpyxl.load_workbook(self.src_path)
        ws_feature_info = self.wb["特性表"]
        feature_info_dict = {}
        for row_info in ws_feature_info.iter_rows(min_row=2):
            if row_info[5].value != None:
                print(row_info[5].value)
                feature_info_dict[row_info[1].value] = row_info[0].value
        return feature_info_dict
    def get_skill_info(self):
        feature_info_dict = self.get_feature_info()
        ws_skill_info = self.wb["check_info"]
        feature_list = list(feature_info_dict.keys())
        for row_index,row_info in enumerate(ws_skill_info.iter_rows()):
            # print(row_index,row_info)
            col_num = 5
            # print(row_info[3].value)
            if row_info[3].value == 0 or row_info[3].value == "无":
                pass
            else:
                for feature in feature_list:
                    # print(feature)
                    if feature in row_info[3].value:
                        ws_skill_info.cell(row=row_index+1,column=col_num,value=f"=特性表!f{feature_info_dict[feature]+1}")
                        col_num += 1

        self.wb.save("src/game_excel/card_all1.xlsx")




if __name__ == "__main__":
    check_excel().get_skill_info()
