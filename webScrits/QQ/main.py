import json

import openpyxl
import requests
from datetime import datetime


class qqUID(object):
    def __init__(self):
        pass
    def qqUID(self,st,end):
        timestamp = datetime.timestamp(datetime.now())
        url = f'https://qun.qq.com/cgi-bin/qun_mgr/search_group_members?bkn=1367143062&ts={timestamp}'
        headres = {
            "Content-type":"application/x-www-form-urlencoded",
            "cookie" : "_qimei_uuid42=18905091c161006dde44de15f680f28635ce2eda19; _qimei_fingerprint=524d0cc475f1645c22acfa4b26e407be; _qimei_q36=; _qimei_h38=a7bc64bcde44de15f680f28603000005118905; RK=+1ms26gjOo; ptcz=598488507d70c464fd3af4d778f97bfd25be112aeeeb25659f7c815ba29b2bd6; rv2=80DEFADC38DBDF201F2299184F0674F4392D5853FABF9AFA36; property20=24CAC1B556F90073167C59A31C27A92A2AD802B0CA6041EF25D8BBF19BB791ACD103628FD4A189BA; pgv_pvid=4818071390; pgv_info=ssid=s2145012202; tgw_l7_route=cc87a893bf42f7c03da1d32ce1078016; uin=o0487372748; skey=@iEgH3BB1L; p_uin=o0487372748; pt4_token=ZxL1E*mvgbMn2RB6pzkfiH8-p*MMHyctOsTDvLYRk84_; p_skey=I25o3K0D-d2x0er4UJaLxNXbFf2dZlQLq-FYUv59k50_; traceid=82d8060a00"
            ,"user-agent":"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36"
        }
        payload = {
            "st": st,
            "end": end,
            "sort": 1,
            "gc": 906815950,
            "bkn": 1367143062
        }
        res = requests.post(url = url,data=payload,headers=headres)
        res_json = json.loads(res.text)
        userInfo = {}
        for uidInfo in res_json["mems"]:
            userInfo[uidInfo["uin"]] = uidInfo["nick"]
        return userInfo


if __name__ == "__main__":
    # wb = openpyxl.Workbook()
    # ws = wb.create_sheet('sheet1')
    # for a in range(0,332,10):
    #     userInfo = qqUID().qqUID(st=a,end=a+9)
    #     for key,value in enumerate(userInfo.items()):
    #         # print(key,type(value))
    #         ws.cell(row=key+a+1,column=1,value=value[0])
    #         ws.cell(row=key + a + 1, column=2, value=value[1])
    #     # break
    # wb.save('uid.xlsx')
    abc = str(['123','123']).replace("'","")
    print(f"'{abc}'")