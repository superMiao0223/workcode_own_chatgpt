import json
import os.path
import discord,requests
import yaml
from discord import app_commands
from datetime import datetime,timezone
import random
import openpyxl
import pymysql
from openpyxl.drawing.image import Image
from FeiShuAPI.feishuapi import feishuApi


configPath = 'conf/info.yaml'
with open(configPath, "r", encoding="utf-8-sig") as f:
    yaml_config = yaml.safe_load(f)
yamlDictInfo = "testGuildInfo"
guildId = yaml_config.get(yamlDictInfo)["guildId"]
channelIdList = yaml_config.get(yamlDictInfo)["channelIdList"]
token = yaml_config.get(yamlDictInfo)["token"]
# MY_GUILD = discord.Object(id=1021963259504504842)  # replace with your guild id
MY_GUILD = discord.Object(id=guildId)

class MyClient(discord.Client):
    def __init__(self, *, intents: discord.Intents):
        super().__init__(intents=intents.all())
        self.tree = app_commands.CommandTree(self)
        self.synced = False
    async def setup_hook(self):
        self.tree.copy_global_to(guild=MY_GUILD)
        await self.tree.sync(guild=MY_GUILD)

client = MyClient(intents=discord.Intents.all())


class getChannelInfo(discord.Client):
    async def getMessageList(self,channel):
        messageListAll = []
        if len(channel.threads) == 0:
            async for messageList in channel.archived_threads():
                messageListAll.append(messageList)
        else:
            for messageList1 in channel.threads:
                messageListAll.append(messageList1)
        return messageListAll
    async def getFormInfo(self,channel):
        """
        如果出现报错，或找不到对应的数据，使用下面这个语句
        # async for messageList in channel.archived_threads():
        """
        messageListAll = await self.getMessageList(channel=channel)
        messageInfoAll = []
        for messageList in messageListAll:
            async for message in messageList.history(oldest_first=True):
                messageTitle = messageList.name
                messageContent = message.content
                messagePostUserId = message.author.id
                messagePostUserName = message.author.name
                messagePostTime = message.created_at
                messageInfoAll.append([messagePostUserId,messagePostUserName,messagePostTime,messageTitle,messageContent,message.attachments])
                # print(messageInfoAll)

        # return messageInfoAll
        """
        修改脚本为上传至飞书文档
        这个是导入到本地文档
        await self.dataToExcel(channelid=channel.id, messageInfoAll=messageInfoAll)
        """
        await self.dataToFeishuExcel(channelid=channel.id, messageInfoAll=messageInfoAll)
    async def getTextInfo(self,channel):
        messageInfoAll = []
        async for message in channel.history(limit=10000):
            messageContent = message.content
            messagePostUserId = message.author.id
            messagePostUserName = message.author.name
            messagePostTime = message.created_at
            messageInfoAll.append([messagePostUserId, messagePostUserName,messagePostTime, "", messageContent, message.attachments])
        print(messageInfoAll)
        """
        修改脚本为上传至飞书文档
        这个是导入到本地文档
        await self.dataToExcel(channelid=channel.id, messageInfoAll=messageInfoAll)
        """
        await self.dataToFeishuExcel(channelid=channel.id, messageInfoAll=messageInfoAll)

    async def dataToExcel(self,messageInfoAll,channelid):
        excelPath = 'src/'+str(datetime.strftime(datetime.today(),"%Y-%m-%d"))+'.xlsx'
        if os.path.exists(excelPath):
            wb = openpyxl.load_workbook(excelPath)
        else:
            wb = openpyxl.Workbook()
        ws = wb.create_sheet(str(channelid))
        rowmax = 2
        valueTitleList = ["发贴人ID","发帖人昵称","发帖时间","主贴标题","帖子内容","帖子图片/附件"]
        for columnNum,valueTitle in enumerate(valueTitleList):
            ws.cell(row=1,column=columnNum+1,value = valueTitle)
        for messageInfoList in messageInfoAll:
            print(messageInfoList)
            if len(messageInfoList[5])>0:
                ws.row_dimensions[rowmax].height = 120
            ws.cell(row=rowmax, column=1, value=str(messageInfoList[0]))
            ws.cell(row=rowmax, column=2, value=messageInfoList[1])
            ws.cell(row=rowmax, column=3, value=datetime.strftime(messageInfoList[2],'%Y-%m-%d'))
            ws.cell(row=rowmax, column=4, value=messageInfoList[3])
            ws.cell(row=rowmax, column=5, value=messageInfoList[4])
            # print(message.attachments)
            messageattsnameList= []
            for messageatts in messageInfoList[5]:
                if messageatts.filename[-4:] == '.png':
                    messageattsname = str(messageatts.id) + '.png'
                    await messageatts.save('image/'+messageattsname)
                    messageattsnameList.append(messageattsname)
            for col,imagepath in enumerate(messageattsnameList):
                imgPath = Image('image/'+imagepath)
                imgPath.width = 60
                imgPath.height = 120
                ws.column_dimensions[str(chr(65 + col + 5))].width = 60
                # imgPath.anchor = imgPath.
                ws.add_image(imgPath,str(chr(65+col+5))+str(rowmax))
            rowmax +=1
        wb.save(excelPath)
        await self.removeImage(messageInfoAll=messageInfoAll)
    async def removeImage(self,messageInfoAll):
        for messageInfoList in messageInfoAll:
            if len(messageInfoList[5])>0:
                for messageatts in messageInfoList[5]:
                    os.remove('image/'+str(messageatts.id) + '.png')

    async def dataToFeishuExcel(self,messageInfoAll,channelid):
        feishuApi().addDataInExcel(dataInfo=messageInfoAll)


@client.event
async def on_ready():
    print(f'Logged in as {client.user} (ID: {client.user.id})')
    print('------')
    if not os.path.exists('image'):
        os.mkdir('image')
    for channelId in channelIdList:
        channelInfo = await client.fetch_channel(int(channelId))
        getInfo = getChannelInfo(intents=discord.Intents.all())
        if str(channelInfo.type) == 'text':
            await getInfo.getTextInfo(channel = channelInfo)
        elif str(channelInfo.type) == 'forum':
            await getInfo.getFormInfo(channel = channelInfo)
        else:
            print(f"其他频道类型：{channelInfo.type}:{channelInfo.id}")
    print("已完成频道内容输出，请查看src文件夹中文件")

#自己机器人
client.run(token)
