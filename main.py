from datetime import datetime
import json,os

import openpyxl
from retrying import retry
# fileList = {
#     'download':['mjwy_dl_day_gp.json','mjwy_dl_day_ios.json'],
#     'revenve':['mjwy_revenve_day_gp.json','mjwy_revenve_day_ios.json']
# }
#
# wb = openpyxl.Workbook()
# ws = wb.active
# for filetype,filepathL in fileList.items():
#     for filepath in filepathL:
#         with open(filepath,'r',encoding='utf-8') as fp:
#             json_str = fp.readline()
#             json_dict = json.loads(json_str)
#             for facts in json_dict["data"]["facets"]:
#                 datemonth = datetime.strftime(datetime.fromtimestamp(facts["date"] / 1000), '%Y-%m')
#                 dateday = datetime.strftime(datetime.fromtimestamp(facts["date"]/1000),'%Y-%m-%d')
#                 device_code = facts["device_code"]
#                 num = facts.get("est_download__sum",facts.get("est_revenue__sum",None))
#                 ws_rowmax = ws.max_row
#                 ws.cell(row=ws_rowmax + 1, column=1, value=datemonth)
#                 ws.cell(row = ws_rowmax+1,column = 2 , value = dateday)
#                 ws.cell(row=ws_rowmax + 1, column=3, value=device_code)
#                 ws.cell(row=ws_rowmax + 1, column=4, value=filetype)
#                 ws.cell(row=ws_rowmax + 1, column=5, value=num)
# wb.save('mjwyData1.xlsx')
@retry
def test1():
    for i in range(0,100):
        print(i)
        if i ==1:
            raise "cuowu"



if __name__ == "__main__":
    # test1()
    time1 = 1728696733
    time2 = 1728610333
    time11 = datetime.date(datetime.fromtimestamp(time1))
    time22 = datetime.date(datetime.fromtimestamp(time2))
    print(time11)
    print(time22)
    if time11 >time22:
        print(1)
    else:
        print(2)