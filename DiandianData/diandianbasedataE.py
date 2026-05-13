import datetime

import openpyxl
from Mymodule.ConnectMysql import Connctmysql as useMysql

class exceldata(object):

    def __init__(self):
        pass

    def toMysql(self):
        wb = openpyxl.load_workbook('src/Squad Busters.xlsx')
        sheetType = {
            "day": ["download","revenue","activeUser"],
            "month" :["activeUser-M","用户留存-全球"]
        }
        for dateType , sheetList in sheetType.items():
            for sheetName in sheetList:
                ws = wb[sheetName]
                if sheetName != "用户留存-全球" :
                    sheetName1 = sheetName.replace("-M","")
                    for i in ws.iter_rows(min_row=2):
                        if i[3].value >0:
                            whereDict1 = {
                                "productName":"Squad Busters",
                                "productID":"1668983788",
                                "date": i[0].value+"/01" if dateType == "month" else i[0].value,
                                "dateType":dateType,
                                "country":i[1].value,
                                "os":"AppStore"
                            }
                            setDict1 = {
                                sheetName1:i[3].value
                            }
                            useMysql(conncetdict={"database": "leiting"}).insert_data(table_name="diandiandatainfo",set_dict=setDict1,where_dict=whereDict1)
                        if i[4].value >0:
                            whereDict2 = {
                                "productName": "Squad Busters",
                                "productID": "com.supercell.squad",
                                "date": i[0].value+"/01" if dateType == "month" else i[0].value,
                                "dateType": dateType,
                                "country": i[1].value,
                                "os": "GooglePlay"
                            }
                            setDict2 = {
                                sheetName1: i[4].value
                            }
                            useMysql(conncetdict={"database": "leiting"}).insert_data(table_name="diandiandatainfo",set_dict=setDict2,where_dict=whereDict2)
                else:
                    for j in ws.iter_rows(min_row=2):
                        whereDict3 = {
                            "productName": "Squad Busters",
                            "productID": j[0].value,
                            "date": str(j[4].value)+"/01",
                            "dateType": dateType,
                            "country": j[3].value,
                            "os": str(j[2].value).replace(" ","")
                        }
                        setDict3 = {
                            "ren2": j[5].value,
                            "ren8": j[6].value,
                            "ren15": j[7].value,
                            "ren31": j[8].value,
                            "ren61": j[9].value,
                            "ren91": j[10].value
                        }
                        useMysql(conncetdict={"database": "leiting"}).insert_data(table_name="diandiandatainfo",set_dict=setDict3,where_dict=whereDict3)



if __name__ == "__main__":
    # exceldata().toMysql()
    time1 = datetime.time(hour=12, minute=5, tzinfo=datetime.timezone.utc)
    print(time1)